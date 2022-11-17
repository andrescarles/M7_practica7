import openpyxl 
  
def escribirexcel():
    wb = openpyxl.Workbook()  #abre conexion con  archivo
    sheet = wb.active  # activa una hoja sobre la que escribir
    c3 = sheet['A1']  # seleccionas la celda
    c3.value = input("valor 1: ")  #introduces valor
    c4 = sheet['B1'] 
    c4.value = input("valor 2: ")
    wb.save("archivo2.xlsx") #guardamos y aqui se crea el archivo.

def leerexcel():
    wb = openpyxl.load_workbook('archivo2.xlsx') #wb es un objeto que contiene el archivo archivo2
    sheet = wb.active #activa la hoja desde la que se desea trabajar
    max_col = sheet.max_column #indicamos el numero maximo de columnas que contiene el archivo ( 2 en este caso)
    
    for i in range(1, max_col + 1): #bucle que por cada columna imprimira el valor de su celda (SOLO LLEGARA A LA ALTURA DE 1 ROW)
        cell_obj = sheet.cell(row = 1, column = i) #Se le especifica que solo recorra una fila por cada columna
        print(cell_obj.value)

    
escribirexcel()
leerexcel()